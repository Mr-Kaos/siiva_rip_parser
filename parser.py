# from datetime import datetime
import os
import re
import fandom_extract
import requests
import time
import mysql.connector
import json
import csv
import traceback
from urllib.parse import unquote
from openpyxl import Workbook, load_workbook

# The csv file of jokes can have multiple search criteria matches if a term is very generic.
# If this is required, separate search terms with the following delimiter: "&|".

wb: Workbook = load_workbook(filename='SiIvaGunnerRips.xlsx',read_only=True)
conn = mysql.connector.connect(
	host="localhost",
	user="root",
	password="",
	database="RipDB"
)

# Executes a stored procedure. Assumes the last parameter is an output parameter and returns it.
def run_sql_proc(proc_name, params):
	proc = conn.cursor()
	result = proc.callproc(proc_name, params)

	return result[len(result) - 1]

def run_sql_query(query, params):
	qry = conn.cursor()
	qry.execute(query, params)
	return qry.fetchall()

class Rip:
	def __init__(self, name, alt_name, desc, upload_date, length, url, yt_id, alt_url, game, channel, genres, jokes, rippers):
		self.name = name
		self.alt_name = alt_name
		self.desc = desc
		self.upload_date = upload_date
		self.length = length
		self.url = url
		self.yt_id = yt_id
		self.alt_url = alt_url
		self.game = game
		self.channel = channel
		self.genres = genres
		self.jokes = jokes
		self.rippers = rippers

		self.commit_to_db()

	def commit_to_db(self):
		# Check if the rip already exists.
		result = run_sql_query("SELECT RipID FROM Rips WHERE RipName = %s", [self.name])

		# If it does, only update its associated jokes, genres and rippers.
		if (len(result) > 0):
			rip_id = result[0][0]

			# Delete and re-insert joke associations
			if self.jokes is not None:
				run_sql_query("DELETE FROM RipJokes WHERE RipID = %s", [rip_id])
				jokes = json.loads(self.jokes)
				for joke in jokes:
					run_sql_query("INSERT INTO RipJokes (RipID, JokeID, JokeTimestamps) VALUES (%s, %s, %s)", [rip_id, joke, json.dumps(jokes[joke]['timestamps'])])

			# Delete and re-insert genre associations
			if self.genres is not None:
				run_sql_query("DELETE FROM RipGenres WHERE RipID = %s", [rip_id])
				for genre in json.loads(self.genres):
					run_sql_query("INSERT INTO RipGenres (RipID, GenreID) VALUES (%s, %s)", [rip_id, genre])

			# Delete and re-insert ripper associations
			if self.rippers is not None:
				run_sql_query("DELETE FROM RipRippers WHERE RipID = %s", [rip_id])
				rippers = json.loads(self.rippers)
				for ripper in rippers:
					run_sql_query("INSERT INTO RipRippers (RipID, RipperID, Alias) VALUES (%s, %s, %s)", [rip_id, ripper, rippers[ripper]])
			
		# Else, insert it
		else:
			run_sql_proc('usp_InsertRip', (self.name, self.alt_name, self.desc, self.upload_date, self.length, self.url, self.yt_id, self.alt_url, self.game, self.channel, self.genres, self.jokes, self.rippers))


	# def debug():
	# 	print(self.name, self.alt_name, self.desc, self.upload, self.length, self.url, self.game, self.channel, self.genres, self.jokes, self.rippers)

def write_missing_joke(joke_data):
	file = open('missing_jokes.txt', 'a')
	file.write(joke_data)
	file.close()

# parses a length field in the spreadsheet and formats it to conform to a MySQL timestamp.
def parse_length(length):
	length = length[2:]
	out = ''

	hour = re.search(r'([0-9]{1,2})H', length)
	if (not hour is None):
		hour = hour.group(1)
		if (len(hour) == 1):
			hour = '0' + hour
		out += hour

	min = re.search(r'([0-9]{1,2})M', length)
	if (not min is None):
		min = min.group(1)
		if (len(min) == 1):
			min = '0' + min
		out += min
	else:
		out += '00'
		
	sec = re.search(r'([0-9]{1,2})S', length)
	if (not sec is None):
		sec = sec.group(1)
		if (len(sec) == 1):
			sec = '0' + sec
		out += sec
	else:
		out += '00'

	return out

def check_genre(genre_name):
	genre_id = 1 # uncategorised
	valid_genres = [
		(2, 'mashup'),
		(2, 'mashups'),
		(3, 'ytp'),
		(3, 'youtube poop'),
		(4, 'remix'),
		(5, 'original'),
		(5, 'original compositions'),
		(6, 'melody swap'),
		(7, 'blue balls'),
		(8, 'arrangement'),
		(9, 'midi'),
		(10, 'medley'),
		(12, 'cover'),
		(12, 'covers'),
		(13, 'sentence mixing'),
		(13, 'sentence mix'),
		(13, 'sentence'),
		(14, 'pitch shifting'),
		(14, 'pitch shift'),
		(14, 'pitch-shifting'),
		(14, 'pitch-shift')
	]

	for genre in valid_genres:
		if genre[1] in genre_name.lower():
			genre_id = genre[0]
			break

	# if (genre_id == 1 and not 'Rips featuring ' in genre_name):
	# 	print('Unknown genre. Given genre: ' + genre_name)
	
	return genre_id

# Looks for the given search in the list of meta jokes
# Search is not case insensitive so that names of songs that are basic words aren't always caught as false matches. (e.g. "Alone" won't match with "all alone")
# This isn't perfect, but will reduce a few false matches.
def check_joke_data(jokes, search):
	actual_joke = None
	actual_meta_jokes = ''
	actual_meta = None
	tags = []
	search = search
	for joke in jokes:
		search_items = joke['joke_search'].split('&|')
		matches = 0
		for criteria in search_items:
			if (criteria in search):
				matches += 1
		if matches == len(search_items):
			actual_joke = joke['joke'].strip()
			actual_meta_jokes = joke['meta_joke'].strip().split('&|')
			actual_meta = joke['meta'].strip()
			tags = joke['tags']
				

	if actual_meta == '':
		actual_meta = None
	if len(actual_meta_jokes) == 0:
		actual_meta_jokes = None
	if (len(tags) == 0):
		tags = None

	return actual_joke, actual_meta_jokes, actual_meta, tags

# Prepares the joke-related data
def prepare_joke_struct(joke, existing_jokes, data):
	new_joke_name, meta_jokes, meta, tags = check_joke_data(existing_jokes, data)
	if joke is None:
		joke = {'name': None, 'timestamps': [], 'meta_joke': None, 'tags': None, 'primary_tag': None}

	if new_joke_name is not None:
		joke['name'] = new_joke_name
		joke['meta_joke_names'] = meta_jokes
		joke['meta_name'] = meta
		meta_id = run_sql_proc('usp_InsertMeta', (meta, 0))
		meta_joke_ids = []
		for mj in meta_jokes:
			meta_joke_ids.append(run_sql_proc('usp_InsertMetaJoke', (mj, None, meta_id, 0)))
		joke['meta_jokes'] = meta_joke_ids
		joke['meta_id'] = meta_id
		joke['primary_tag'] = None
		joke['tags'] = None
		if (tags is not None):
			# If the first tag is not empty, set it as primary
			if (tags[0] != ''):
				tag_id = run_sql_proc('usp_InsertTag', (tags[0], 0))
				joke['primary_tag'] = tag_id
			# If there are other tags, add them (if they are also not empty)
			if (len(tags) > 1):
				joke['tags'] = []
				for tag in tags[1:]:
					if (tag != ''):
						tag_id = run_sql_proc('usp_InsertTag', (tag, 0))
						joke['tags'].append(tag_id)

				if len(joke['tags']) == 0:
					joke['tags'] = None

	return joke

# Inserts or updates the given joke into the database, and returns its ID.
def insert_joke(joke):
	# print(json.dumps(joke))
	meta_jokes = None
	if joke['meta_joke'] is not None:
		meta_jokes = json.dumps([joke['meta_joke']])
	tags = None
	if joke['tags'] is not None:
		tags = json.dumps(joke['tags'])

	joke_id = 0

	# If the joke exists, update its metas
	result = run_sql_query("SELECT JokeID FROM Jokes WHERE JokeName = %s", [joke['name']])
	if (len(result) > 0):
		joke_id = result[0][0]
		run_sql_query("DELETE FROM JokeMetas WHERE JokeID = %s", [joke_id])
		for meta_id in joke['meta_jokes']:
			run_sql_query("INSERT INTO JokeMetas (JokeID, MetaJokeID) VALUES (%s, %s)", [joke_id, meta_id])
	# else, insert it
	else:
		joke_id = run_sql_proc('usp_InsertJoke', (joke['name'], None, joke['primary_tag'], tags, meta_jokes, 0))
	return joke_id

# validates the given timestamp to a format supported by the database
def validate_timestamp(timestamp):
	validated = None

	segments = timestamp.split(':')
	parsed = [];
	for seg in segments:
		if len(seg) < 2:
			seg = '0' + seg
		parsed.append(seg)

	validated = ':'.join(parsed)
	if len(segments) < 3:
		validated = '00:' + validated

	return validated

def parse_jokes(text_box, existing_jokes):
	# Get jokes segment
	joke_ids = {}

	# First check if a jokes table exists. If it does, use it instead.
	matches = re.search(r'==.*Joke', text_box, re.IGNORECASE)
	# if matches is None:
	# 	matches = re.search(r'==.*setlist', text_box, re.IGNORECASE)
	if not matches is None:
		jokes_start = matches.start()
		table_start = text_box.find('{|', jokes_start)

		# If no table:
		if table_start == -1:
			heading_end = text_box.find('==', jokes_start + 3)
			jokes_end = text_box.find('==', heading_end + 3)
			joke_data = text_box[heading_end+3:jokes_end-1].strip()

			joke = prepare_joke_struct(None, existing_jokes, joke_data)
			if joke['name'] is not None:
				joke_id = insert_joke(joke)
				joke_ids[joke_id] = {'timestamps': joke['timestamps'], 'comment': None}
			# else:
			# 	print('no joke match!')
				# write_missing_joke(joke_data)

		# Else, check the contents of the jokes table
		else:
			table_start = text_box.find('{|');
			table_start = text_box.find('|-', table_start)
			jokes_end = text_box.find('|}\n', table_start)
			joke_data = text_box[table_start+3:jokes_end+2]

			# parse jokes table. Read 4 lines at a time.
			last_start = None
			last_end = None
			skip_timestamp_rows = 0
			invalid_joke = False
			joke = {'name': None, 'timestamps': [], 'meta_joke': None, 'tags': None, 'primary_tag': None}
			debug_prev_line = None

			lineCounter = 0
			for line in joke_data.splitlines():
				# print (line, lineCounter, skip_timestamp_rows)
				if len(line) > 0:

					# If the line is a header, ignore it.
					if (line[0] != '!'):

						# If end of joke, Add data to list
						if (line == '|-' or line == '|}'):
							if skip_timestamp_rows > 0:
								skip_timestamp_rows -= 1

							# If the joke is not invalid, add it. Disable this if statement to add invalid jokes.
							if not joke['name'] is None and not invalid_joke:
								joke_id = insert_joke(joke)
								joke_ids[joke_id] = {'timestamps': joke['timestamps'], 'comment': None}

							# reset joke
							joke = {'name': None, 'timestamps': [], 'meta_joke': None, 'tags': None, 'primary_tag': None}
							invalid_joke = False
							if (skip_timestamp_rows > 0):
								lineCounter = 1
							else:
								lineCounter = 0
							continue

						# If the first line of a joke, but the timestamp is omitted, set the line counter to the next line to get the joke name
						if lineCounter == 0:
							# if 'rowspan' in line:
							# 	skip_timestamp_rows = int(re.search(r'rowspan?.=?.([0-9])', line).groups()[0])

							if skip_timestamp_rows == 0:
								# print('parsing timestamp:', line, skip_timestamp_rows)
								# If the timestamp has a rowspan, skip a row

								# If the timestamp has a start and end, get it
								if ('-' in line):
									timestamps = re.search(r'(?:[0-9]{0,2}:?[0-9]{0,2}:[0-9]{0,2}).-.(?:[0-9]{0,2}:?[0-9]{0,2}:[0-9]{0,2})', line)
									if (timestamps is not None and timestamps.group() is not None):
										time = timestamps.group()
										split = time.split('-')
										start = validate_timestamp(split[0].strip())
										end = validate_timestamp(split[1].strip())
										joke['timestamps'].append({'start': start, 'end': end})
										last_start = start
										last_end = end
								else:
									timestamps = re.search(r'(?:[0-9]{0,2}:?[0-9]{0,2}:[0-9]{0,2})', line)
									last_end = None
									if timestamps is not None:
										if len(timestamps.groups()) > 0:
											for time in timestamps.groups():
												time = validate_timestamp(time)
												joke['timestamps'].append({'start': time})
												last_start = time
										else:
											last_start = validate_timestamp(timestamps.group())
											joke['timestamps'].append({'start': last_start})
							else:
								print (last_start, last_end)
								if last_start is not None and last_end is None:
									joke['timestamps'].append({'start': last_start})
								elif last_end is not None:
									joke['timestamps'].append({'start': last_start, 'end': last_end})

							if 'rowspan' in line:
								r = re.search(r'rowspan?.=?.([0-9])', line)
								if r is not None:
									skip_timestamp_rows = int(r.groups()[0])

						# joke name
						elif lineCounter == 1:
							if ('???' in line):
								invalid_joke = True

							line = line[1:].replace("''", '').strip()
							if (len(line) > 0):
								
								# If joke name starts with quotation marks, remove them from start and end.
								if (line[0] == '"'):
									line = line[1:len(line) -1]

								joke_name = None
								matches = re.search(r'\[\[(.*)\]\]', line, re.IGNORECASE)
								if (matches is None):
									joke_name = line.strip()
								else:
									joke_name = matches.groups()[0]

								joke = prepare_joke_struct(joke, existing_jokes, joke_name)
							else:
								invalid_joke = True

					lineCounter += 1
					debug_prev_line = line

	return joke_ids
	
# Returns a dictionary of jokes, meta jokes, their metas and tags of the joke.
def read_joke_metas(meta_file):
	jokes = []
	with open(meta_file, 'r') as csv_file:
		reader = csv.reader(csv_file)
		next(reader, None)

		for row in reader:
			jokes.append({
			'joke': row[0],
			'joke_search': row[1],
			'meta_joke': row[2],
			'meta': row[3],
			'tags': row[4:]
			})

	return jokes

# Checks if the given fandom page has been cached for quicker retrieval. If not, caches the page.
def load_fandom_page(folder, url):
	page_data = None
	cached = False
	decoded = unquote(url)
	decoded = decoded.replace('/', '')
	decoded = decoded.replace('https:', '')
	decoded = decoded.replace('.fandom.comwiki', '_')
	file_name = "./cached/" + folder + '/' + decoded + ".txt"

	# Make sure the cache dir exists
	if not os.path.exists("./cached/" + folder):
		os.mkdir("./cached/" + folder)

	# If the file does not exist, extract the page from fandom and save it.
	if not os.path.exists(file_name) and not os.path.exists(decoded):
		file = open(file_name, "w")

		response = requests.get(url + '?action=edit')
		parser = fandom_extract._build_parser()
		args = parser.parse_args()
		page_data = fandom_extract.parse_wtsource(response.content, page_type=args.page_type)
		file.write(page_data)
		file.close()
	else:
		file = open(file_name, "r")
		page_data = file.read()
		cached = True

	return page_data, cached

def find_all_matches(pattern, string, group=0):
    pat = re.compile(pattern)
    pos = 0
    out = []
    while m := pat.search(string, pos):
        pos = m.start() + 1
        out.append(m[group])
    return out

def get_fandom_data(text_box, metas, game):
	game = None
	track = None
	alt_url = None
	categories = None
	rippers = None
	jokes = None

	# print (url, text_box)

	# Game name
	# Only used to check if the wiki page is valid.
	match = re.search(r'playlist*.=(.*)', text_box)
	if (match is not None):
		if (len(match.groups()) > 0):
			game = match.group(1).strip()
			# If the rip is an announcement, skip it.
			if ('announcements;' in game.lower()):
				game = None
			else:
				# In case there is a formatting error and the table row (| character) is not on a new line, find it and end the text there.
				cutoff = game.find('|')
				if cutoff != -1:
					game = game[:cutoff]

				if len(game) <= 256:
					game = run_sql_proc('usp_InsertGame', (game, None, 0))
				else:
					game = None

	if (not game is None):
		# Track name
		match = re.search(r'track*.= *.\[(http.*?) (.*?)\]', text_box)
		if (match is not None):
			if (len(match.groups()) > 0):
				alt_url = match.group(1).strip()
				track = match.group(2).strip()

		# Genres
		categories = []
		for cat in re.findall(r'^|.*Category*.:(.*)]]', text_box):
			if (cat != ''):
				cat = check_genre(cat)
				if cat not in categories:
					categories.append(cat)

		# If a category other than 1 exists, remove it.
		if (len(categories) > 1 and 1 in categories):
			categories.remove(1)

		categories = json.dumps(categories)

		# Rippers
		rippers = {}
		substring = '|author'
		start_pos = text_box.find(substring)
		if start_pos == -1:
			substring = '| author'
			start_pos = text_box.find(substring)
		# If the author field is still not found, do not attempt to get it.
		if (start_pos == -1):
			print('Cannot find ripper!')
		else:
			ripper_text = text_box[start_pos + len(substring):text_box.find('\n', start_pos)].strip()
			# remove the '=' at the beginning
			ripper_text = ripper_text[1:]

			# If '<ref>' tags exist, remove them
			if ('<ref>' in ripper_text):
				ripper_text = ripper_text[:ripper_text.find('<ref>')].strip()
			# Else, match for double square braces.
			# (Exclude cases where it states "see [[x", as these are usually links to a another page)
			if ('see [[' not in ripper_text) and ('#' not in ripper_text) and ('<!--' not in ripper_text):
				# matches = re.search(r'(|.*author*=*.)\[\[(.*?)\]\]', ripper_text)
				matches = re.finditer(r'(|.*author*=*.)\[\[(.*?)\]\]', ripper_text)
				if (matches is not None):
					matchFound = False

					# If matches are found, add them all
					for match in matches:
						matchFound = True
						ripper = match.groups()[1]
						alias = None
						if ('|' in ripper):
							split = ripper.split('|')
							ripper = split[0]
							alias = split[1]

						ripper = run_sql_proc('usp_InsertRipper', (ripper, 0))
						rippers[ripper] = alias

					# If no match was found, add the line
					if not matchFound and ripper_text != '':
						# In case there is a formatting error and the table row (| character) is not on a new line, find it and end the text there.
						cutoff = ripper_text.find('|')
						if cutoff != -1:
							ripper_text = ripper_text[:cutoff]
						
						ripper = run_sql_proc('usp_InsertRipper', (ripper_text.strip(), 0))
						rippers[ripper] = None
				else:
					ripper_text = ripper_text.strip()
					if (ripper_text != ''):
						ripper = run_sql_proc('usp_InsertRipper', (ripper_text.strip(), 0))
						rippers[ripper] = None
			
			# if len(rippers) == 0:
			# 	print ("No rippers found!")
		rippers = json.dumps(rippers)

		# Jokes
		jokes = parse_jokes(text_box, metas)
		if jokes is not None:
			jokes = json.dumps(jokes)
		
	# Game is not returned as it often produces unreliable results.
	return track, alt_url, categories, jokes, rippers, game

#parses a worksheet from the spreadsheet and adds it to the array of rips.
def parse_worksheet(sheet_name, channel, metas):
	ws = wb[sheet_name]
	row_start = 0
	rowNum = row_start
	rows = 40000

	for row in ws.iter_rows(min_row=2 + row_start, max_col=6,max_row=row_start + rows + 1,values_only=True):
		# print(rowNum)
		name_search = re.search(r', "(.*)"', row[1])
		if name_search is None:
			print ("Malformed cell! Row:", rowNum)
		else:
			name = re.search(r', "(.*)"', row[1]).group(1)
			url = re.search(r'^.*\("(.*)?",', row[0]).group(1)
			date = row[4]
			yt_id = re.search(r'"([A-Za-z0-9_\-]{11})"', row[0])
			if (yt_id is not None):
				yt_id = yt_id.group(1)
			else:
				yt_id = None

			# If the length is empty, skip rip
			if row[5] != None:
				length = parse_length(row[5])

				try:
					fandom_url = re.search(r'^.*\("(.*)?",', row[1]).group(1)
					# print(name, fandom_url)
					text_box, cached = load_fandom_page(sheet_name, fandom_url)

					if (text_box == ''):
						print('Null wiki page: "' + fandom_url + '"')
					else:
						track, alt_url, genres, jokes, rippers, alt_game = get_fandom_data(text_box, metas, game = None)

						game = re.search(r'", ".* [-](.*)"', row[1])
						if (game is not None):
							game = game.group(1).strip()
							game = run_sql_proc('usp_InsertGame', (game, None, 0))

						# If the game is not obtained from the spreadsheet, use the name from the fandom page.
						if game is None:
							game = alt_game

						# print(name, track, None, date, length, url, alt_url, game, channel, genres, jokes, rippers)

						if not [x for x in (name, date, length, url, game) if x is None]:
							Rip(name, track, None, date, length, url, yt_id, alt_url, game, channel, genres, jokes, rippers)
						else:
							print('Cannot insert rip as a required value is null! - ' + fandom_url)
							print(name, date, length, url, game)

						# 0.5 second delay to avoid overloading fandom with requests
						if not cached:
							time.sleep(1.0)
				except Exception as exc:
					print ('Failed to obtain rip data. Row: ' + str(rowNum) + ' fandom: ' + fandom_url)
					traceback.print_tb(exc.__traceback__)
					traceback.print_exc()

		rowNum += 1

# Read joke data
file = open('joke_sample.txt', 'r')
jokes = read_joke_metas('meta_jokes.csv')

parse_worksheet('SiIvaGunner', 2, jokes)
parse_worksheet('TimmyTurnersGrandDad', 3, jokes)

